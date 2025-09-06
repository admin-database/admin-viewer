# -*- coding: utf-8 -*-
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def autosize_excel(path: str):
    wb = load_workbook(path)
    ws = wb.active
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = ""
            try:
                v = str(cell.value) if cell.value is not None else ""
            except:
                pass
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
    ws.freeze_panes = "A2"
    wb.save(path)

def parse_id_list(text: str):
    if not text:
        return []
    parts = [p.strip() for p in re.split(r"[,\s]+", text) if p.strip()]
    return list(dict.fromkeys(parts))

def sanitize_component(s: str) -> str:
    if not s: return "값없음"
    s = re.sub(r'[\\/:*?"<>|]+', "_", str(s))
    s = re.sub(r"\s+", "_", s.strip())
    return s or "값없음"

def is_url(s: str) -> bool:
    s = (s or "").strip().lower()
    return s.startswith("http://") or s.startswith("https://")
